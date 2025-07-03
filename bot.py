import os
import random
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters  # type: ignore
from quizzes_data import quizzes
from openpyxl import Workbook, load_workbook

user_state = {}

def log_to_excel(name, action, subject=None, lecture=None, score=None, total=None):
    file = "user_logs.xlsx"
    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        ws.append(["Full Name", "Action", "Subject", "Lecture", "Score", "Total", "Date"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb["Logs"]
    date = datetime.now().strftime("%Y-%m-%d")
    ws.append([
        name,
        action,
        subject if subject else "",
        lecture if lecture else "",
        score if score is not None else "",
        total if total is not None else "",
        date
    ])
    wb.save(file)

def get_subjects():
    return sorted([name for name in os.listdir("lectures") if os.path.isdir(f"lectures/{name}")])

def get_types(subject):
    path = f"lectures/{subject}"
    folders = [name for name in os.listdir(path) if os.path.isdir(f"{path}/{name}")]
    return sorted(folders) if folders else [""]

def get_lectures(subject, type_):
    path = f"lectures/{subject}/{type_}" if type_ else f"lectures/{subject}"
    return sorted([name for name in os.listdir(path) if name.endswith(".pdf")])

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user_state[uid] = {}
    keyboard = [[s] for s in get_subjects()] + [["🏠 القائمة الرئيسية"]]
    await update.message.reply_text("📚 اختر المادة:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    name = update.effective_user.full_name
    text = update.message.text
    state = user_state.get(uid, {})

    if text == "🏠 القائمة الرئيسية":
        return await start(update, context)

    if text == "🔙 الرجوع للخلف":
        if "lecture" in state:
            del user_state[uid]["lecture"]
            lectures = get_lectures(state["subject"], state["type"])
            keyboard = [[l] for l in lectures] + [["🏠 القائمة الرئيسية"]]
            await update.message.reply_text("📖 اختر المحاضرة:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))
        elif "type" in state:
            del user_state[uid]["type"]
            types = get_types(state["subject"])
            keyboard = [[t] for t in types] + [["🏠 القائمة الرئيسية"]]
            await update.message.reply_text("📘 اختر النوع :", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))
        elif "subject" in state:
            return await start(update, context)
        return

    if text in get_subjects():
        user_state[uid] = {"subject": text}
        if text == "Adults":
            keyboard = [
                ["🧪 امتحان شامل", "📚  المحاضرات النظري وكويزات خفيفة"],
                ["أسئلة الدكتورة"],
                ["🏠 القائمة الرئيسية"]
            ]
            await update.message.reply_text("📘 اختر نوع المحتوى:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))
        else:
            types = get_types(text)
            if types != [""]:
                keyboard = [[t] for t in types] + [["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
                await update.message.reply_text("📘 اختر النوع :", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))
            else:
                lectures = get_lectures(text, "")
                keyboard = [[l] for l in lectures] + [["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
                user_state[uid]["type"] = ""
                await update.message.reply_text("📖 اختر المحاضرة:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

    elif "subject" in state and text == "📚  المحاضرات النظري وكويزات خفيفة":
        subject = state["subject"]
        types = get_types(subject)
        if types != [""]:
            keyboard = [[t] for t in types] + [["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
            await update.message.reply_text("📘 اختر النوع :", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))
        else:
            lectures = get_lectures(subject, "")
            keyboard = [[l] for l in lectures] + [["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
            user_state[uid]["type"] = ""
            await update.message.reply_text("📖 اختر المحاضرة:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

    elif "subject" in state and text in ["🧪 امتحان شامل", "أسئلة الدكتورة"]:
        lecture_key = {
            "🧪 امتحان شامل": "exam",
            "أسئلة الدكتورة": "Question_Bank"
        }[text]

        if lecture_key not in quizzes:
            await update.message.reply_text("❗ لا توجد أسئلة متاحة لهذا القسم حالياً.")
            return

        mcqs = quizzes[lecture_key].get("MCQs", [])
        tfs = quizzes[lecture_key].get("TF", [])
        random.shuffle(mcqs)
        random.shuffle(tfs)

        user_state[uid]["quiz"] = {
            "lecture": lecture_key,
            "current": 0,
            "score": 0,
            "mcqs": mcqs,
            "tfs": tfs
        }

        question_data = mcqs[0]
        keyboard = [[opt] for opt in question_data["options"]] + [["⛔️ إنهاء الكويز"], ["🏠 القائمة الرئيسية"]]
        await update.message.reply_text(
            f"🧪 السؤال 1:\n{question_data['question']}",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        )

    elif "subject" in state and text in get_types(state["subject"]):
        user_state[uid]["type"] = text
        lectures = get_lectures(state["subject"], text)
        keyboard = [[l] for l in lectures] + [["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
        await update.message.reply_text("📖 اختر المحاضرة:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

    elif "subject" in state and "type" in state and text in get_lectures(state["subject"], state["type"]):
        user_state[uid]["lecture"] = text.replace(".pdf", "").strip()
        keyboard = [["📄 View Lecture File", "📝 Take Quiz"], ["🏠 القائمة الرئيسية"], ["🔙 الرجوع للخلف"]]
        await update.message.reply_text(f"📘 {text}\nاختر ما تريد:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True))

    elif text == "📄 View Lecture File":
        subject = state.get("subject")
        type_ = state.get("type")
        lecture = state.get("lecture") + ".pdf"
        file_path = f"lectures/{subject}/{type_}/{lecture}" if type_ else f"lectures/{subject}/{lecture}"
        if os.path.exists(file_path):
            with open(file_path, "rb") as f:
                await update.message.reply_document(f)
        else:
            await update.message.reply_text("❌ الملف غير موجود.")
        log_to_excel(name, "Viewed Lecture", subject, state.get("lecture"))

    elif text == "📝 Take Quiz":
        lecture = state.get("lecture", "").strip()
        if not lecture or lecture not in quizzes:
            await update.message.reply_text("❗ لا يوجد كويز مضاف لهذه المحاضرة حتى الآن.")
            return

        mcqs = quizzes[lecture].get("MCQs", [])
        tfs = quizzes[lecture].get("TF", [])
        random.shuffle(mcqs)
        random.shuffle(tfs)

        user_state[uid]["quiz"] = {
            "lecture": lecture,
            "current": 0,
            "score": 0,
            "mcqs": mcqs,
            "tfs": tfs
        }

        question_data = mcqs[0]
        keyboard = [[opt] for opt in question_data["options"]] + [["⛔️ إنهاء الكويز"], ["🏠 القائمة الرئيسية"]]
        await update.message.reply_text(
            f"🧪 السؤال 1:\n{question_data['question']}",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        )

    elif text == "⛔️ إنهاء الكويز" and "quiz" in state:
        quiz = state["quiz"]
        score = quiz["score"]
        total = len(quiz["mcqs"]) + len(quiz["tfs"])
        log_to_excel(name, "Quiz Manual End", state.get("subject"), state.get("lecture"), score, total)
        del user_state[uid]["quiz"]
        await update.message.reply_text(f"⛔️ الكويز تم إنهاؤه يدويًا.\n✅ درجتك: {score}/{total}")
        return

    elif "quiz" in state:
        quiz = state["quiz"]
        mcqs = quiz.get("mcqs", [])
        tfs = quiz.get("tfs", [])
        total_mcq = len(mcqs)
        total_tf = len(tfs)
        total_all = total_mcq + total_tf
        current = quiz["current"]

        if current < total_mcq:
            q = mcqs[current]
            correct_letter = q["answer"].split(".")[0].strip().upper()
            chosen = text[0].upper()
            if chosen == correct_letter:
                quiz["score"] += 1
                feedback = "✅ إجابة صحيحة!"
            else:
                correct_text = next((opt for opt in q["options"] if opt.strip().upper().startswith(correct_letter)), q["answer"])
                feedback = f"❌ إجابة خاطئة.\n✅ الإجابة الصحيحة: {correct_text}"

        elif current < total_all:
            tf_index = current - total_mcq
            q = tfs[tf_index]
            correct_answer = q["answer"]
            chosen_answer = True if "true" in text.lower() else False
            if chosen_answer == correct_answer:
                quiz["score"] += 1
                feedback = "✅ إجابة صحيحة!"
            else:
                feedback = f"❌ إجابة خاطئة.\n✅ الإجابة الصحيحة: {'True' if correct_answer else 'False'}"

        quiz["current"] += 1
        current += 1

        if current < total_all:
            await update.message.reply_text(feedback)
            if current < total_mcq:
                q = mcqs[current]
                keyboard = [[opt] for opt in q["options"]] + [["⛔️ إنهاء الكويز"], ["🏠 القائمة الرئيسية"]]
            else:
                q = tfs[current - total_mcq]
                keyboard = [["✅ True"], ["❌ False"], ["⛔️ إنهاء الكويز"], ["🏠 القائمة الرئيسية"]]
            await update.message.reply_text(
                f"🧪 {q['question']}",
                reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
            )
        else:
            await update.message.reply_text(feedback)
            score = quiz["score"]
            log_to_excel(name, "Quiz Finished", state.get("subject"), state.get("lecture"), score, total_all)
            del user_state[uid]["quiz"]
            await update.message.reply_text(f"✅ انتهى الكويز!\nدرجتك: {score}/{total_all}")
    else:
        await update.message.reply_text("❗ من فضلك اختر من القوائم.", reply_markup=ReplyKeyboardMarkup([["🏠 القائمة الرئيسية"]], resize_keyboard=True))

# ✅ تشغيل البوت
app = ApplicationBuilder().token("7774771769:AAHXK9PVehCzEh5d9NOksBlD4UyfqbZ5ObM").build()
app.add_handler(CommandHandler("start", start))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

print("🤖 Bot is running... Send /start to begin.")
app.run_polling()
